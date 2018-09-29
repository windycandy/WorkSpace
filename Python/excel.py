
import openpyxl
from openpyxl import load_workbook


#workbook = xlrd.open_workbook(r'C:\Users\user\Desktop\test.xls')
workbook = load_workbook(r'C:\Users\user\Desktop\test.xlsx')
print(workbook.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = workbook.get_sheet_by_name("Sheet1")

#print(sheet["A"])
#print(sheet["4"])
print(sheet["C4"].value)

print(sheet.max_row)   
print(sheet.max_column) 

for i in sheet["C"]:
  print(i.value, end = "/") 

for i in range(10):
  sheet["A%d" % (i+1)].value = '000000'
  
workbook.save(r'C:\Users\user\Desktop\test.xlsx')