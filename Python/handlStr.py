
import openpyxl
from openpyxl import load_workbook

h_f = 3  #begin at line 3

workbook_old = load_workbook(r'/home/ckt/work2/sp/stringTable_old.xlsx')
#print(workbook_from.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
sheet_o = workbook_old.get_sheet_by_name("All")
print(sheet_o.max_row)   
print(sheet_o.max_column) 


workbook_n = load_workbook(r'/home/ckt/work2/sp/stringTable_new.xlsx')
#print(workbook_to.sheetnames)    
sheet_n = workbook_n.get_sheet_by_name("itel")
print(sheet_n.max_row)   
print(sheet_n.max_column)



for i in range(sheet_o.max_row - 2):
	isFind = 0
	id_o = sheet_o["E%d" % (i+2)].value
	
	if id_o != None:
		if sheet_o["D%d" % (i+2)].value != "false" or sheet_o["C%d" % (i+2)].value != "bool":
			for j in range(sheet_n.max_row - 2):
				if sheet_n["B%d" % (j+2)].value == id_o:
					print("E%d" % (i+2) , "   ",  "B%d" % (j+2) )
					
					if sheet_n["M%d" % (j+2)].value != None:
						sheet_o["Q%d" % (i+2)].value = sheet_n["M%d" % (j+2)].value
					if sheet_n["O%d" % (j+2)].value != None:	
						sheet_o["S%d" % (i+2)].value = sheet_n["O%d" % (j+2)].value
					


  


workbook_old.save(r'/home/ckt/work2/sp/stringTable_0525_PRO.xlsx')
